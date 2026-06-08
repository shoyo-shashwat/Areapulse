"""
AreaPulse Portal — export_engine.py
PDF (reportlab) and Excel (openpyxl) export.
Degrades gracefully if libraries not installed.
"""
import io
import time
from datetime import datetime

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.units import cm
    _REPORTLAB = True
except ImportError:
    _REPORTLAB = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


def export_pdf_summary(issues, report_type='summary', officer_name='', dept=''):
    """
    Generate a PDF report. Returns bytes or None.
    report_type: 'summary' | 'sla' | 'category' | 'area'
    """
    if not _REPORTLAB:
        return None

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=2*cm, bottomMargin=2*cm,
        leftMargin=2*cm, rightMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    story  = []

    # Header
    header_style = ParagraphStyle('Header', parent=styles['Heading1'],
                                  fontSize=20, textColor=colors.HexColor('#1E3A5F'),
                                  spaceAfter=6)
    sub_style    = ParagraphStyle('Sub', parent=styles['Normal'],
                                  fontSize=10, textColor=colors.HexColor('#8A7060'))
    body_style   = ParagraphStyle('Body', parent=styles['Normal'],
                                  fontSize=10, textColor=colors.HexColor('#1A1208'),
                                  leading=14)

    story.append(Paragraph('AreaPulse Civic Intelligence Report', header_style))
    story.append(Paragraph(
        f'Generated: {datetime.now().strftime("%d %b %Y, %I:%M %p IST")} · '
        f'Officer: {officer_name or "System"} · Department: {dept or "All"}',
        sub_style
    ))
    story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#DED8CC')))
    story.append(Spacer(1, 0.4*cm))

    # Summary stats
    open_c     = sum(1 for i in issues if i.get('status') != 'resolved')
    resolved_c = sum(1 for i in issues if i.get('status') == 'resolved')
    breached_c = sum(1 for i in issues if i.get('sla_state') == 'breached')

    stats_data = [
        ['Metric', 'Count'],
        ['Total Issues', str(len(issues))],
        ['Open / Active', str(open_c)],
        ['Resolved', str(resolved_c)],
        ['SLA Breached', str(breached_c)],
    ]
    stats_table = Table(stats_data, colWidths=[9*cm, 4*cm])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F7F5F0'), colors.white]),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#DED8CC')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
    ]))
    story.append(stats_table)
    story.append(Spacer(1, 0.5*cm))

    # Issue table
    story.append(Paragraph('Issue Details', ParagraphStyle('H2', parent=styles['Heading2'],
                                                            fontSize=13, textColor=colors.HexColor('#1E3A5F'))))
    story.append(Spacer(1, 0.2*cm))

    table_data = [['#AP', 'Area', 'Category', 'Severity', 'Status', 'SLA', 'Description']]
    for i in issues[:100]:  # cap at 100 rows
        desc = (i.get('description') or '')[:50]
        sla  = i.get('sla_state', 'healthy').upper().replace('_', ' ')
        table_data.append([
            str(i.get('id', '')),
            i.get('area', ''),
            (i.get('tag') or '').title(),
            (i.get('severity') or '').title(),
            (i.get('status') or '').replace('_', ' ').title(),
            sla,
            desc,
        ])

    col_widths = [1.2*cm, 2.5*cm, 2*cm, 1.8*cm, 2.2*cm, 2*cm, 5.3*cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F7F5F0'), colors.white]),
        ('GRID',       (0, 0), (-1, -1), 0.3, colors.HexColor('#DED8CC')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
    ]))

    # Color SLA-breached rows red
    for row_idx, issue in enumerate(issues[:100], start=1):
        if issue.get('sla_state') == 'breached':
            t.setStyle(TableStyle([('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#FAE4E2'))]))

    story.append(t)

    # Footer
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        f'AreaPulse GovNGO Portal · Confidential Government Document · {datetime.now().year}',
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8,
                       textColor=colors.HexColor('#B8A890'), alignment=1)
    ))

    doc.build(story)
    buf.seek(0)
    return buf.read()


def export_excel(issues, filename='areapulse-issues'):
    """
    Generate an Excel file. Returns bytes or None.
    """
    if not _OPENPYXL:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Issues'

    header_fill   = PatternFill('solid', fgColor='1E3A5F')
    header_font   = Font(color='FFFFFF', bold=True, size=11)
    breached_fill = PatternFill('solid', fgColor='FAE4E2')
    center        = Alignment(horizontal='center', vertical='center')
    thin          = Side(style='thin', color='DED8CC')
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['Ticket', 'Area', 'Category', 'Description', 'Severity', 'Status',
               'SLA State', 'Upvotes', 'Filed At', 'Assigned To']
    ws.append(headers)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = center
        cell.border = border

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20

    for issue in issues:
        ts = issue.get('timestamp', 0)
        filed = datetime.fromtimestamp(ts).strftime('%d %b %Y %H:%M') if ts else ''
        row = [
            f"AP-{issue.get('id', '')}",
            issue.get('area', ''),
            (issue.get('tag') or '').title(),
            (issue.get('description') or '')[:120],
            (issue.get('severity') or '').title(),
            (issue.get('status') or '').replace('_', ' ').title(),
            (issue.get('sla_state') or 'healthy').replace('_', ' ').upper(),
            issue.get('upvotes', 0),
            filed,
            issue.get('assigned_to') or '',
        ]
        ws.append(row)
        row_idx = ws.max_row
        for col in range(1, 11):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            if issue.get('sla_state') == 'breached':
                cell.fill = breached_fill

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
